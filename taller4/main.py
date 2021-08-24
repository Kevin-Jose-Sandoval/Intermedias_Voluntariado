import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
import string

def automate_excel(file_name):
    ''' input -> sales_mes.xlsx | output -> report_mes.xlsx'''
    # Read the excel file
    file = pd.read_excel(file_name)

    # index -> how to classify, columns = Product line and the values= = Total
    table = file.pivot_table( index = 'Gender', 
                            columns = 'Product line', 
                            values = 'Total', 
                            aggfunc = 'sum' ).round(0)

    month_extension = file_name.split('_')[1]
    
    # start default column
    table.to_excel(f'report_{month_extension}', startrow = 4, sheet_name = 'Report')

    # working with openpyxl
    workbook = load_workbook(f'report_{month_extension}')
    tab = workbook['Report']

    # identify reference columns (workbook)
    # minimum column: A(1) and maximum column: G(7)
    # minimum row: 5 and maximum row: 7
    min_column = workbook.active.min_column
    max_column = workbook.active.max_column
    min_row = workbook.active.min_row
    max_row = workbook.active.max_row

    # make the graph 
    barchart = BarChart()

    data = Reference(tab, 
                    min_col = min_column + 1, # Gender column we don't need it
                    max_col = max_column, 
                    min_row = min_row,
                    max_row = max_row
                    )

    categories = Reference(tab, 
                    min_col = min_column,
                    max_col = min_column, 
                    min_row = min_row + 1,  # header row we don't need it
                    max_row = max_row
                    )

    # titles_from_data = True : headers included
    barchart.add_data(data, titles_from_data = True)
    barchart.set_categories(categories)

    # Add graphic
    # B12: where does it start 
    tab.add_chart(barchart, 'B12')
    barchart.title = 'Sales'
    barchart.style = 2

    # formulas
    alphabet = list(string.ascii_uppercase)
    alphabet_excel = alphabet[0:max_column]

    for i in alphabet_excel:
        if i!='A':
            tab[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})' # tab['B8'] = '=SUM(B6:B7)'
            tab[f'{i}{max_row+1}'].style = 'Currency'
            
    tab[f'{alphabet_excel[0]}{max_row+1}'] = 'Total'

    tab['A1'] = 'Reporte'
    month = month_extension.split('.')[0]
    tab['A2'] = month
    tab['A1'].font = Font('Arial', bold=True, size=20)
    tab['A2'].font = Font('Arial', bold=True, size=12)

    workbook.save(f'report_{month_extension}')
    
def main():
    try:
        automate_excel('sales_2021.xlsx')
        print("Proceso terminado correctamene")
    except:
        print("Ocurrió un error en el proceso")
    


if __name__ == '__main__':
    main()