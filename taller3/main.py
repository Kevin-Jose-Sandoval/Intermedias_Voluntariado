import openpyxl

# Open the excel document 
book = openpyxl.load_workbook(filename='data.xlsx')
sheet = book['clientes']

# empty dictionary 
customers = {}

# range from row 2 until to the last
for i in range(2, sheet.max_row + 1):
    # get the column one
    customer = sheet.cell( row = i, column = 1 ).value
     # invoice amount 
    amount = sheet.cell( row = i, column = 4 ).value 
    # if key is not in dict yet
    customers.setdefault(customer, 0)
    # if exist key
    customers[customer] += amount


# range from row 2 until to the last
for i in range(2, sheet.max_row + 1):
    # get the column one
    customer = sheet.cell( row = i, column = 1 ).value
     # total column
    total = sheet.cell( row = i, column = 5 )

    total.value = customers[customer]

book.save('output.xlsx')